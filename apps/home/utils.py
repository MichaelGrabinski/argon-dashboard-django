# apps/home/utils.py

import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from io import BytesIO

def export_to_excel(filename, columns, data_rows):
    """
    - filename: 'something.xlsx'
    - columns: list of tuples (header, key) where key may be 'field' or 'related__field'
    - data_rows: QuerySet of model instances or list of dicts
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.replace('.xlsx', '')

    # Write headers
    header_font = Font(bold=True)
    for idx, (header, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, (_, key) in enumerate(columns, start=1):
            if hasattr(row, '__class__') and not isinstance(row, dict):
                # row is model instance
                if '__' in key:
                    parts = key.split('__')
                    val = row
                    for p in parts:
                        val = getattr(val, p, '')
                        if val is None:
                            val = ''
                            break
                else:
                    val = getattr(row, key, '')
            else:
                # row is dict
                val = row.get(key, '')

            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto‐adjust column widths (simple approach)
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    # Save to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
